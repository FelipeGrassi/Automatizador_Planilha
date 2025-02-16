import openpyxl
from tkinter import *
import tkinter as tk
from tkinter import ttk

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

coluna_empresa=[]
coluna_colaborador=[]
coluna_ocupação=[]
coluna_exame=[]
letra=["A","B","C","D","E","G"]
i=1
medicos=["Dr.Guilherme Paniguel","Dra.Sonia Maria de Barros","Dra.Thaina"]
exames=["Exame Clínico","Audiometria","Acuidade","Eletrocardiograma","Eletroencefalograma","Espirometria",
        "Raio-x(Torax)","Raio-x(Lombar)","Avaliação Psicossocial","Avaliação Epworth","Teste Romberg",
        "Hemograma","Glicemia","GamaGT",'TGO',"TGP","Colinesterase Plasmática","Colinesterase Eritrócitaria",
        "Acido Hipurico","Acido Metil Hipurico","RastIGE","Parasitológico de Fezes","Coprocultura","Micológico de Unha",
        "VDRL","Cultura de Fezes","Urina","Reticulócitos","Toxicológico"]
ocupacional=["Adm","Dem","Per","Rt","Mro","Ass"]

def next():
    px_clbd=puxar_colaborador.get()    #atribui os dados
    px_emps=puxar_empresa.get()

    exa = lista_exame.curselection()  # utiliza o curselection para retorna indices selecionados
    lista_ex = [lista_exame.get(n) for n in exa]  # realiza a operação dentro de uma lista para puxar os exames do indices
    lis = "+".join(lista_ex)

    ocu=ocupacional.get()

    coluna_empresa.append(px_emps)
    coluna_colaborador.append(px_clbd)
    coluna_ocupação.append(ocu)
    coluna_exame.append(lis)





    puxar_colaborador.delete(0,tk.END)    #reseta a linha para preencher com novos dados
    puxar_empresa.delete(0, tk.END)


def Salvar_Excell():
    tipo=".xlsx"
    data=nome.get()
    data_excell=data+tipo


    a="A"
    b="B"
    c="C"
    d="D"
    e="E"
    f="F"
    g="G"
    contador = 7
    rere="Audiometria"



    wb = openpyxl.Workbook()
    ws=wb.active

    #cabeçalho
    ws["A6"]="N"
    ws.column_dimensions['A'].width=3
    ws.row_dimensions[6].height=19
    ws.column_dimensions['B'].width =11
    ws.column_dimensions['C'].width = 33
    ws.column_dimensions['D'].width = 33
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 95
    ws.column_dimensions['G'].width = 10
    fonte_cab = Font(name="Arial Black", size=11)
    cores=PatternFill(fill_type="solid", fgColor="848484")
    bordas=Border(left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"))



    ws["B6"]="Data"
    ws["C6"]="Empresa"
    ws["D6"]="Colaborador"
    ws["E6"] = "Exames"
    ws.merge_cells("E6:F6")
    ws["G6"]="Audio"
    ws["A6"].font = fonte_cab
    ws["B6"].font = fonte_cab
    ws["C6"].font = fonte_cab
    ws["D6"].font = fonte_cab
    ws["E6"].font = fonte_cab
    ws["G6"].font = fonte_cab
    ws["G5"]="=SOMA(G7:G50)"
    ws["E6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A6"].border=bordas
    ws["B6"].border = bordas
    ws["C6"].border = bordas
    ws["D6"].border = bordas
    ws["E6"].border = bordas
    ws["F6"].border = bordas
    ws["G6"].border = bordas
    ws["A6"].fill = cores
    ws["B6"].fill=cores
    ws["C6"].fill = cores
    ws["D6"].fill = cores
    ws["E6"].fill = cores
    ws["F6"].fill = cores
    ws["G6"].fill = cores
    medicodia=medico.get()
    ws["A1"]=medicodia



    for i ,(em,cos,ocu,ex) in enumerate(zip(coluna_empresa,coluna_colaborador,coluna_ocupação,coluna_exame),start=1):

        co=str(contador)
        posa=a+co
        posb= b+co
        posc=   c + co
        posd = d + co
        pose = e + co
        posf = f + co
        posg = g + co
        ws[posa]=i
        ws[posb]=data
        ws[posc]=em
        ws[posd] = cos
        ws[pose] = ocu
        ws[posf] = ex
        ws[posf].alignment = Alignment(horizontal="center", vertical="center")

        contador=contador+1
        i=i+1
        if rere in ex:
            ws[posg]=1








    wb.save(data_excell)





x=tk.Tk()
x.title("Planilha Fechamento")
x.config(bg="#87CEEB")
x.geometry("600x500")

#nome da planilha
nome=Entry()
nome.place(x=250,y=5)
nome.insert(0,"Data")


#puxar dados da empre
empresa=Label(x,text="Digite a empresa :",bg="#6E6E6E",fg="#E6E6E6").place(x=5,y=30)
puxar_empresa=Entry(x,width=50)
puxar_empresa.place(x=105,y=30)

#puxa o nome do colaborador
colaborador=Label(x,text="Colaborador:",bg="#6E6E6E",fg="#E6E6E6").place(x=30,y=60)
puxar_colaborador=Entry(x,width=50)
puxar_colaborador.place(x=105,y=60)

#lista exames
lista_exame=tk.Listbox(x,bg="#6E6E6E",width=100,height=50,selectmode=EXTENDED)
lista_exame.place(x=415,y=30)
for lista in exames:
    lista_exame.insert(tk.END,lista)

#tipo de exame ocupacional
ocupacional=ttk.Combobox(x,values=ocupacional,width=20)
ocupacional.current(0)
ocupacional.place(x=100,y=90)

#Lista de Medicos
medico=ttk.Combobox(x,values=medicos,width=38)
medico.current(0)
medico.place(x=10,y=350)



B1=Button(x,bg="#6E6E6E",text="Proximo",command=next).place(x=330,y=120)
B2=Button(x,bg="#6E6E6E",text="Salvar Planilha",command=Salvar_Excell).place(x=290,y=350)


x.mainloop()